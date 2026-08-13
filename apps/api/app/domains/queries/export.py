"""Export helpers: CSV and XLSX with spreadsheet-formula injection protection.

A cell whose value starts with =, +, -, @ or a tab is prefixed with a single
quote so spreadsheet apps treat it as text, never as a formula. Persian text,
RTL sheets, dates and amounts are formatted where the format permits.
"""

from __future__ import annotations

import csv
import datetime as dt
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DANGEROUS_PREFIXES = ("=", "+", "-", "@", "\t")


def _safe(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, str):
        if value.startswith(DANGEROUS_PREFIXES):
            return "'" + value
        return value
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    return value


def _header_row(columns: list[dict[str, object]]) -> list[str]:
    return [str(c["label"]) for c in columns]


def to_csv(columns: list[dict[str, object]], rows: list[list[object]]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_header_row(columns))
    for row in rows:
        writer.writerow([_safe(cell) for cell in row])
    return buf.getvalue()


def to_csv_bytes(columns: list[dict[str, object]], rows: list[list[object]]) -> bytes:
    return ("\ufeff" + to_csv(columns, rows)).encode("utf-8")  # BOM for Excel


def to_xlsx_bytes(columns: list[dict[str, object]], rows: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "گزارش"
    ws.sheet_view.rightToLeft = True  # RTL sheet

    header_fill = PatternFill("solid", fgColor="14604F")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, label in enumerate(_header_row(columns), start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_safe(value))  # type: ignore[call-overload]

    # column widths (approx)
    for idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
